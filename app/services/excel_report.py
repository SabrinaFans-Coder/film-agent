import json
from io import BytesIO
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
ASSISTANT_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws, ncols):
    for col in range(1, ncols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)


async def export_session_excel(session_id: str) -> bytes:
    """Generate xlsx for a single session's chat history."""
    from app.api.routes import agent_core
    from fastapi import HTTPException

    try:
        history = await agent_core.get_history(session_id)
    except HTTPException:
        history = None

    wb = Workbook()
    ws = wb.active
    ws.title = "对话记录"

    headers = ["角色", "内容", "时间"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 3)

    if history is not None:
        for i, msg in enumerate(history.messages, 2):
            ws.cell(row=i, column=1, value=msg.role)
            ws.cell(row=i, column=2, value=msg.content)
            ws.cell(row=i, column=3, value=str(msg.timestamp))
            if msg.role == "assistant":
                for col in range(1, 4):
                    ws.cell(row=i, column=col).fill = ASSISTANT_FILL

    _auto_width(ws, 3)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_stats_excel() -> bytes:
    """Generate global statistics xlsx with three sheets."""
    from app.db.checkpoint import list_sessions_meta

    sessions, _ = list_sessions_meta(limit=10000, offset=0)

    wb = Workbook()

    # Sheet 1: 影片推荐
    ws1 = wb.active
    ws1.title = "影片推荐"
    headers1 = ["片名", "出现次数", "TMDB评分", "类型"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    _style_header(ws1, 4)

    movie_counter: Counter = Counter()
    movie_info = {}
    for s in sessions:
        try:
            movies = json.loads(s.get("movies_mentioned", "[]"))
        except (json.JSONDecodeError, TypeError):
            movies = []
        for m in movies:
            # Skip entries with no structured data (old unenriched or non-movie)
            if not m.get("genres") and m.get("tmdb_rating") is None:
                continue
            title = m.get("title", "未知")
            movie_counter[title] += 1
            if title not in movie_info:
                movie_info[title] = {
                    "rating": m.get("tmdb_rating"),
                    "genres": "、".join(m.get("genres", [])),
                }

    for i, (title, count) in enumerate(movie_counter.most_common(), 2):
        info = movie_info.get(title, {})
        ws1.cell(row=i, column=1, value=title)
        ws1.cell(row=i, column=2, value=count)
        ws1.cell(row=i, column=3, value=info.get("rating"))
        ws1.cell(row=i, column=4, value=info.get("genres", ""))

    _auto_width(ws1, 4)

    # Sheet 2: 类型分布
    ws2 = wb.create_sheet("类型分布")
    headers2 = ["类型", "数量", "占比"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 3)

    genre_counter: Counter = Counter()
    for s in sessions:
        try:
            movies = json.loads(s.get("movies_mentioned", "[]"))
        except (json.JSONDecodeError, TypeError):
            movies = []
        for m in movies:
            for g in m.get("genres", []):
                genre_counter[g] += 1

    total_genres = sum(genre_counter.values()) or 1
    for i, (genre, count) in enumerate(genre_counter.most_common(), 2):
        ws2.cell(row=i, column=1, value=genre)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=f"{count / total_genres:.1%}")

    _auto_width(ws2, 3)

    # Sheet 3: 会话概览
    ws3 = wb.create_sheet("会话概览")
    headers3 = ["会话ID", "消息数", "创建时间", "更新时间"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 4)

    for i, s in enumerate(sessions, 2):
        ws3.cell(row=i, column=1, value=s["session_id"])
        ws3.cell(row=i, column=2, value=s["message_count"])
        ws3.cell(row=i, column=3, value=s["created_at"])
        ws3.cell(row=i, column=4, value=s["updated_at"])

    _auto_width(ws3, 4)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
